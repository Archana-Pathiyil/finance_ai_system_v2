"""
Report Service - Generate downloadable Excel reports
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from services.ap_service import get_payment_due_report, get_vendor_wise_summary
from services.ar_service import get_ageing_report, get_collection_followup
from services.reconciliation_service import get_reconciliation_results


def _style_header_row(ws, row_num: int, num_cols: int, fill_color: str = "1F3864"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generate_ap_outstanding_report(db: Session) -> bytes:
    """Generate AP Outstanding Excel report."""
    data = get_vendor_wise_summary(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Outstanding"

    headers = ["Vendor Name", "Vendor Code", "Invoice Count", "Total Amount (AED)", "Outstanding (AED)", "Currency"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for row in data:
        ws.append([
            row["vendor_name"], row.get("vendor_code", ""),
            row["invoice_count"], row["total_amount"], row["total_outstanding"], row["currency"]
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_payment_due_report(db: Session) -> bytes:
    """Generate Payment Due Schedule Excel report."""
    data = get_payment_due_report(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Due"

    headers = ["Vendor", "Invoice No", "Invoice Date", "Due Date", "Amount", "Outstanding", "Days Until Due", "Status"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers), "C0392B")

    status_fills = {
        "overdue": PatternFill("solid", fgColor="FADBD8"),
        "due": PatternFill("solid", fgColor="FDEBD0"),
        "upcoming": PatternFill("solid", fgColor="D5F5E3"),
    }

    for i, row in enumerate(data, start=2):
        ws.append([
            row["vendor_name"], row["invoice_no"], row["invoice_date"],
            row["due_date"], row["amount"], row["outstanding"],
            row["days_until_due"], row["status"]
        ])
        fill = status_fills.get(row["status"])
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_ar_ageing_report(db: Session) -> bytes:
    """Generate AR Ageing Excel report."""
    data = get_ageing_report(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "AR Ageing"

    headers = ["Customer", "Total Outstanding", "0-30 Days", "31-60 Days", "61-90 Days", "90+ Days"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers), "1A5276")

    for row in data:
        ws.append([
            row["customer_name"], row["total_outstanding"],
            row["bucket_0_30"], row["bucket_31_60"],
            row["bucket_61_90"], row["bucket_90_plus"],
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_collection_followup_report(db: Session) -> bytes:
    """Generate Collection Follow-up Excel report."""
    data = get_collection_followup(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection Followup"

    headers = ["Customer", "Invoice No", "Invoice Date", "Days Outstanding", "Amount", "Outstanding", "Bucket", "Currency"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers), "784212")

    for row in data:
        ws.append([
            row["customer_name"], row["invoice_no"], row["invoice_date"],
            row["days_outstanding"], row["amount"], row["outstanding"],
            row["bucket"], row["currency"]
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_reconciliation_report(db: Session, rec_type: str) -> bytes:
    """Generate Reconciliation Excel report."""
    data = get_reconciliation_results(db, rec_type)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{rec_type.upper()} Reconciliation"

    headers = ["Party Name", "Invoice No", "ERP Amount", "SOA Amount", "Difference", "Status"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers), "1F618D")

    status_fills = {
        "matched": PatternFill("solid", fgColor="D5F5E3"),
        "discrepancy": PatternFill("solid", fgColor="FDEBD0"),
        "unmatched": PatternFill("solid", fgColor="FADBD8"),
    }

    for i, row in enumerate(data, start=2):
        ws.append([
            row["party_name"], row["invoice_no"],
            row["erp_amount"], row["soa_amount"], row["difference"], row["status"]
        ])
        fill = status_fills.get(row["status"])
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_party_statement_ap(db: Session, vendor_name: str) -> bytes:
    """Generate a detailed AP statement for a specific vendor."""
    from models.models import VendorLedger, ReconciliationResult
    from utils.date_utils import calculate_due_date, get_payment_status
    from config import AP_PAYMENT_DAYS
    from datetime import date

    records = db.query(VendorLedger).filter(
        VendorLedger.vendor_name == vendor_name
    ).order_by(VendorLedger.invoice_date).all()

    recon = {r.invoice_no: r for r in db.query(ReconciliationResult).filter(
        ReconciliationResult.reconciliation_type == "ap",
        ReconciliationResult.party_name == vendor_name
    ).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Statement"

    # Header block
    ws.merge_cells("A1:H1")
    ws["A1"] = f"VENDOR STATEMENT — {vendor_name.upper()}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="1A3C6E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Statement Date:"; ws["B2"] = date.today().isoformat()
    ws["A2"].font = Font(bold=True)
    ws["D2"] = "Currency:"; ws["E2"] = "AED"
    ws["D2"].font = Font(bold=True)
    ws.append([])

    headers = ["Invoice No", "Invoice Date", "Due Date", "Invoice Amount (AED)",
               "Paid Amount (AED)", "Outstanding (AED)", "Recon Status", "Fine/Difference (AED)"]
    ws.append(headers)
    _style_header_row(ws, 4, len(headers), "1A3C6E")

    total_invoiced = total_paid = total_outstanding = total_fine = 0.0

    for r in records:
        due = r.due_date or calculate_due_date(r.invoice_date, AP_PAYMENT_DAYS)
        rec = recon.get(r.invoice_no)
        recon_status = rec.status if rec else "Not Reconciled"
        fine = abs(rec.difference or 0) if rec and rec.difference else 0.0

        row = [
            r.invoice_no,
            r.invoice_date.isoformat() if r.invoice_date else "",
            due.isoformat(),
            r.amount, r.paid_amount, r.outstanding or 0,
            recon_status, fine
        ]
        ws.append(row)
        total_invoiced += r.amount
        total_paid += r.paid_amount
        total_outstanding += r.outstanding or 0
        total_fine += fine

    # Totals row
    next_row = ws.max_row + 1
    ws.cell(next_row, 1, "TOTAL").font = Font(bold=True)
    for col, val in [(4, total_invoiced), (5, total_paid), (6, total_outstanding), (8, total_fine)]:
        ws.cell(next_row, col, round(val, 2)).font = Font(bold=True)
        ws.cell(next_row, col).fill = PatternFill("solid", fgColor="F0F4FF")
        ws.cell(next_row, col).number_format = "#,##0.00"

    for col in range(4, 9):
        for row in range(5, next_row):
            ws.cell(row, col).number_format = "#,##0.00"

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_party_statement_ar(db: Session, customer_name: str) -> bytes:
    """Generate a detailed AR statement for a specific customer."""
    from models.models import CustomerLedger, ReconciliationResult
    from utils.date_utils import get_ageing_bucket
    from datetime import date

    records = db.query(CustomerLedger).filter(
        CustomerLedger.customer_name == customer_name
    ).order_by(CustomerLedger.invoice_date).all()

    recon = {r.invoice_no: r for r in db.query(ReconciliationResult).filter(
        ReconciliationResult.reconciliation_type == "ar",
        ReconciliationResult.party_name == customer_name
    ).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Statement"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"CUSTOMER STATEMENT — {customer_name.upper()}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="145A32")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Statement Date:"; ws["B2"] = date.today().isoformat()
    ws["A2"].font = Font(bold=True)
    ws["D2"] = "Currency:"; ws["E2"] = "AED"
    ws["D2"].font = Font(bold=True)
    ws.append([])

    headers = ["Invoice No", "Invoice Date", "Ageing Bucket", "Invoice Amount (AED)",
               "Received Amount (AED)", "Outstanding (AED)", "Recon Status", "Fine/Difference (AED)"]
    ws.append(headers)
    _style_header_row(ws, 4, len(headers), "145A32")

    total_invoiced = total_received = total_outstanding = total_fine = 0.0

    for r in records:
        bucket = get_ageing_bucket(r.invoice_date) if r.invoice_date else "N/A"
        rec = recon.get(r.invoice_no)
        recon_status = rec.status if rec else "Not Reconciled"
        fine = abs(rec.difference or 0) if rec and rec.difference else 0.0

        ws.append([
            r.invoice_no,
            r.invoice_date.isoformat() if r.invoice_date else "",
            bucket,
            r.amount, r.received_amount, r.outstanding or 0,
            recon_status, fine
        ])
        total_invoiced += r.amount
        total_received += r.received_amount
        total_outstanding += r.outstanding or 0
        total_fine += fine

    next_row = ws.max_row + 1
    ws.cell(next_row, 1, "TOTAL").font = Font(bold=True)
    for col, val in [(4, total_invoiced), (5, total_received), (6, total_outstanding), (8, total_fine)]:
        ws.cell(next_row, col, round(val, 2)).font = Font(bold=True)
        ws.cell(next_row, col).fill = PatternFill("solid", fgColor="F0FFF4")
        ws.cell(next_row, col).number_format = "#,##0.00"

    for col in range(4, 9):
        for row in range(5, next_row):
            ws.cell(row, col).number_format = "#,##0.00"

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
